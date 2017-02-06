import openpyxl
from collections import OrderedDict


def write_mapper():
    data = OrderedDict()
    data['pnr'] = ['200101011234', '200202022468']
    data['id_nr'] = [1, 2]
    wb = openpyxl.Workbook()
    sheet = wb.active
    fill_sheet(sheet, data)
    wb.save('testdata/mapper.xlsx')


def fill_sheet(sheet, data):
    for col, (name, values) in enumerate(data.items(), start=1):
        sheet.cell(column=col, row=1).value = name
        for row, value in enumerate(values, start=2):
            sheet.cell(column=col, row=row).value = value


def write_test_data():
    wb = openpyxl.Workbook()
    sheet = wb.active
    data = OrderedDict()
    data['pnr'] = ['200101011234', '200202022468', '200101011234', '200202022468', ]
    data['id_nr'] = ['', '', '', '', ]
    data['visit'] = ['2017-01-01', '2017-02-02', '2017-03-03', '2017-04-04', ]
    fill_sheet(sheet, data)
    wb.create_sheet(title='Weights')
    sheet = wb.get_sheet_by_name('Weights')
    data = OrderedDict()
    data['date'] = ['2017-01-01', '2017-02-02', '2017-03-03', '2017-04-04', '2018-12-12']
    data['id_nr'] = ['', '', '', '', '']
    data['pnr'] = ['200101011234', '200202022468', '200101011234', '200202022468', '202002022222']
    data['weight'] = [34.5, 35.5, 36.6, 37.5, 22.2]
    fill_sheet(sheet, data)
    wb.save('testdata/unmapped_data.xlsx')


if __name__ == '__main__':
    write_mapper()
    write_test_data()
