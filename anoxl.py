from tkinter import Tk, ttk, StringVar, Listbox, N, S, W, E, Scrollbar, VERTICAL, END
from tkinter.filedialog import askopenfilename, asksaveasfilename
import openpyxl


class AnoXL:
    def gui_init(self):
        root = Tk()
        root.title("AnoXL - Thinkware AB")

        self.mapper_path = StringVar()
        self.sensitive_id = StringVar()
        self.anonymous_id = StringVar()

        ttk.Button(
            root,
            text='Select mapping workbook',
            command=self.open_mapping
        ).grid(row=0, column=0)

        ttk.Label(root, text='Mapper file: ').grid(row=1, column=0)
        ttk.Label(root, textvariable=self.mapper_path).grid(row=1, column=1, columnspan=3)

        ttk.Label(root, text='Sensitive id: ').grid(row=2, column=0)
        self.sens_id = ttk.Combobox(root, textvariable=self.sensitive_id)
        self.sens_id.grid(row=2, column=1)
        ttk.Label(root, text='Anonymous id: ').grid(row=2, column=2)
        self.anon_id = ttk.Combobox(root, textvariable=self.anonymous_id)
        self.anon_id.grid(row=2, column=3)

        ttk.Button(
            root,
            text='Select data workbook',
            command=self.open_data
        ).grid(row=3, column=0)

        ttk.Label(root, text='Change log').grid(row=4, column=0)

        self.yScroll = Scrollbar(root, orient=VERTICAL)
        self.yScroll.grid(row=5, column=4, sticky=N + S)

        self.listbox = Listbox(
            root,
            yscrollcommand=self.yScroll.set,
            activestyle='none',
        )
        self.listbox.grid(row=5, column=0, columnspan=4, sticky=N + S + E + W)
        self.yScroll['command'] = self.listbox.yview

        root.mainloop()

    def log(self, text):
        self.listbox.insert(END, text)

    def open_mapping(self):
        mapper_name = askopenfilename(
            initialdir=".",
            filetypes=(("Excel Workbook", "*.xlsx"), ("All Files", "*.*")),
            title="Select Excel file with id mappings."
        )
        self.prepare_mappings(mapper_name)

    def prepare_mappings(self, mapper_name):
        self.mapper_path.set(mapper_name)
        self.mapper = MappingFile(mapper_name)
        self.sens_id['values'] = tuple(self.mapper.names)
        self.anon_id['values'] = tuple(self.mapper.names)

    def open_data(self):
        data_name = askopenfilename(
            initialdir=".",
            filetypes=(("Excel Workbook", "*.xlsx"), ("All Files", "*.*")),
            title="Select Excel file with data to add anonymous id in."
        )
        data_file = DataFile(data_name, self.log)
        data_file.set_sens_id(self.sensitive_id.get(), self.mapper.values(self.sensitive_id.get()))
        data_file.set_anon_id(self.anonymous_id.get(), self.mapper.values(self.anonymous_id.get()))
        data_file.process_sheets()
        self.log('Done')
        path = asksaveasfilename(
            initialdir=".",
            filetypes=(("Excel Workbook", "*.xlsx"), ("All Files", "*.*")),
            title="Store data with anonymous id added here."
        )
        if path:
            data_file.save(path)
            self.log('Saved ' + path)
        else:
            self.log('No file name provided. Nothing saved.')


class MappingFile:
    def __init__(self, filename):
        wb = openpyxl.load_workbook(filename)
        self.mapper_sheet = wb.active
        self.names = []
        for col in range(1, self.mapper_sheet.max_column + 1):
            value = self.mapper_sheet.cell(column=col, row=1).value
            if value:
                self.names.append(value)

    def values(self, column_name):
        vals = []
        for col in range(1, self.mapper_sheet.max_column + 1):
            value = self.mapper_sheet.cell(column=col, row=1).value
            if value == column_name:
                break
        else:
            return []
        for row in range(1, self.mapper_sheet.max_row + 1):
            value = self.mapper_sheet.cell(column=col, row=row).value
            if value:
                vals.append(value)
        return vals


class DataFile:
    def __init__(self, filename, log):
        self.wb = openpyxl.load_workbook(filename)
        self.log = log
        log('Opened ' + filename)

    def set_sens_id(self, name, ids):
        self.sens_id_name = name
        self.sens_id_list = ids

    def set_anon_id(self, name, ids):
        self.anon_id_name = name
        self.anon_id_list = ids

    def process_sheets(self):
        self.mapping = dict(zip(self.sens_id_list, self.anon_id_list))
        for sheet_name in self.wb.get_sheet_names():
            self.log('Processing sheet ' + sheet_name)
            ws = self.wb.get_sheet_by_name(sheet_name)
            self.process_sheet(ws)

    def process_sheet(self, ws):
        n = 0
        anon_ix = 0
        sens_ix = 0
        for col in range(1, ws.max_column + 1):
            value = ws.cell(column=col, row=1).value
            if value == self.anon_id_name:
                anon_ix = col
            elif value == self.sens_id_name:
                sens_ix = col
        if not anon_ix:
            self.log('No column for anonymous id found. Skipping sheet.')
            return
        if not sens_ix:
            self.log('No column for sensitive id found. Skipping sheet.')
            return
        for row in range(2, ws.max_row + 1):
            sens_id = ws.cell(row=row, column=sens_ix).value
            anon_id = self.mapping.get(sens_id)
            if anon_id:
                ws.cell(row=row, column=anon_ix).value = anon_id
                n += 1
            else:
                self.log('No anonymous id found for ' + sens_id)
        self.log('Processed {} rows.'.format(n))

    def save(self, path):
        self.wb.save(path)


if __name__ == '__main__':
    AnoXL().gui_init()
